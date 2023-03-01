import pandas as pd
from plotly.offline import plot
import numpy as np
import plotly.offline as py
import plotly.graph_objs as go
import plotly.express as px
import tempfile
import os
import matplotlib.pyplot as plt
import calendar
import graficos
import datetime
import locale
import openpyxl
from openpyxl import workbook 
from openpyxl import load_workbook
import io
import zipfile
from flask import Flask, send_file, make_response

def convert_to_time(decimal_num):
    hours = int(decimal_num)
    minutes = int(round((decimal_num - hours) * 60))
    return f"{hours:02d}:{minutes:02d}"



def preenche_modelo(mes,ano,nome): #ex: 1,Presencial
	grupos=['Grupo de Pesquisa','Programa']
	tipo=['Presencial','Remoto']
	df         = pd.read_excel('atividade.xlsx')
	df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
	df         = df.loc[df['DATA'].dt.month == mes]
	df         = df.loc[df['DATA'].dt.year == ano]
	df['Hora'] = df['HORAS'].apply(lambda x: x.hour + x.minute / 60 + x.second / 3600)
	locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
	month_name = datetime.date(2000, mes, 1).strftime('%B')
	month_name = month_name.capitalize()
	wb = load_workbook('modelo.xlsx')
	sheets = wb.sheetnames
	Sheet1 = wb[sheets[0]] ##
	Sheet1.cell(row = 3, column = 1).value = nome
	Sheet1.cell(row = 3, column = 2).value = month_name
	Sheet1.cell(row = 3, column = 3).value = ano
	total= df['Hora'].sum() 
	Sheet1.cell(row = 10, column = 2).value = convert_to_time(total)
	Sheet1.cell(row = 7, column = 2).value = convert_to_time(total)
	df_sum     = df.groupby(['GRUPO']).agg({'Hora': 'sum'}).reset_index()
	df_sum2    = df.groupby(['tipo']).agg({'Hora': 'sum'}).reset_index()
	for index, (n, i) in enumerate(zip(grupos, tipo)):
		horas = df_sum.loc[df_sum['GRUPO'] == n, 'Hora'].iloc[0]
		Sheet1.cell(row = 8 + index, column = 2).value = convert_to_time(horas)
		Sheet1.cell(row = 8 + index, column = 3).value = horas*100/total
		horas = df_sum2.loc[df_sum2['tipo'] == i, 'Hora'].iloc[0]
		Sheet1.cell(row = 5 + index, column = 2).value = convert_to_time(horas)
		Sheet1.cell(row = 5 + index, column = 3).value = horas*100/total
	df_sum3    = df.groupby(['SUBCATEGORIA']).agg({'Hora': 'sum'}).reset_index()
	for index, row in df_sum3.iterrows():
		Sheet1.cell(row = 12 + index, column = 1).value = row['SUBCATEGORIA']
		Sheet1.cell(row = 12 + index, column = 2).value = convert_to_time(row['Hora'])
		df4    = df[df['SUBCATEGORIA']==row['SUBCATEGORIA']]
		Sheet1.cell(row = 12 + index, column = 3).value = df4['ATIVIDADE'].nunique()
	Sheet1.title = 'Relatorio'
	df = df.drop(['Hora'], axis=1)
	df['DATA'] = df['DATA'].dt.strftime('%d/%m/%Y') 
	df['HORAS'] = df['HORAS'].dt.strftime('%H:%M')
	
	df.to_excel('temp.xlsx', sheet_name='Dados detalhados',index=False)
	temp_wb = openpyxl.load_workbook('temp.xlsx')
	ws2 = wb.create_sheet("Dados detalhados")
	for row in temp_wb['Dados detalhados']:
		for cell in row:
			ws2[cell.coordinate].value = cell.value
	new_file_name = 'relatorio_'+month_name+'.xlsx'
	wb.save(new_file_name)
	os.remove('temp.xlsx')
	return new_file_name
	'''
	planilha = pd.ExcelWriter('relatorio_'+month_name+'.xlsx', engine='openpyxl',mode='w')
	planilha.book = wb
	df.to_excel(planilha, sheet_name='Dados detalhados',index=False)
	return planilha
	'''

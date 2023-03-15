import dash
import dash_html_components as html
import dash_table
import dash_core_components as dcc
from   dash.dependencies import Input, Output, State
import pandas as pd
from   datetime import datetime
import datetime
import base64
import locale
import graficos
import zipfile
import io
import matplotlib.pyplot as plt
from   plotly.offline import plot
import numpy as np
import plotly.offline as offline
import plotly.graph_objs as go
import plotly.express as px
import tempfile
import os
import calendar
import graficos
import openpyxl
from   openpyxl import workbook 
from   openpyxl import load_workbook
from   flask import Flask, send_file, make_response
import requests
import colorlover as cl

def convert_to_time(decimal_num):
    hours = int(decimal_num)
    minutes = int(round((decimal_num - hours) * 60))
    return f"{hours:02d}:{minutes:02d}"

def retorna_df(contents, filename):
	contents=str(contents[0])
	content_type, content_string = contents.split(',')
	decoded = base64.b64decode(content_string)
	df = pd.read_excel(io.BytesIO(decoded))
	return df
# Criando um DataFrame com dados de exemplo
def preenche_modelo(mes,ano,nome,df): #ex: 1,Presencial
	url = 'https://github.com/Grenda07/ppgtu/blob/main/src/modelo.xlsx?raw=true'
	response = requests.get(url)
	content = response.content
	file = io.BytesIO(content)
	grupos=['Grupo de Pesquisa','Programa']
	tipo=['Presencial','Remoto']
	df['DATA'] = pd.to_datetime(df['DATA'], dayfirst=True)
	df         = df.loc[df['DATA'].dt.month == mes]
	df         = df.loc[df['DATA'].dt.year == ano]
	df['Hora'] = df['HORAS'].apply(lambda x: x.hour + x.minute / 60 + x.second / 3600)
	locale.setlocale(locale.LC_TIME, 'pt_BR.UTF-8')
	month_name = datetime.date(2000, mes, 1).strftime('%B')
	month_name = month_name.capitalize()
	wb = load_workbook(file)
	'''
	sheets = wb.sheetnames
	Sheet1 = wb[sheets[0]] ##
	Sheet1.cell(row = 3, column = 1).value = nome
	Sheet1.cell(row = 3, column = 2).value = month_name
	Sheet1.cell(row = 3, column = 3).value = ano
	total= df['Hora'].sum() 
	Sheet1.cell(row = 10, column = 2).value = convert_to_time(total)
	Sheet1.cell(row = 7, column = 2).value = convert_to_time(total)
	df_sum     = df.groupby(['GRUPO']).agg({'Hora': 'sum'}).reset_index()
	df_sum2    = df.groupby(['tipo']).agg({'Hora': 'sum'}).reset_index()
	for index, (n, i) in enumerate(zip(grupos, tipo)):
		horas = df_sum.loc[df_sum['GRUPO'] == n, 'Hora'].iloc[0]
		Sheet1.cell(row = 8 + index, column = 2).value = convert_to_time(horas)
		Sheet1.cell(row = 8 + index, column = 3).value = horas*100/total
		horas = df_sum2.loc[df_sum2['tipo'] == i, 'Hora'].iloc[0]
		Sheet1.cell(row = 5 + index, column = 2).value = convert_to_time(horas)
		Sheet1.cell(row = 5 + index, column = 3).value = horas*100/total
	df_sum3    = df.groupby(['SUBCATEGORIA']).agg({'Hora': 'sum'}).reset_index()
	for index, row in df_sum3.iterrows():
		Sheet1.cell(row = 12 + index, column = 1).value = row['SUBCATEGORIA']
		Sheet1.cell(row = 12 + index, column = 2).value = convert_to_time(row['Hora'])
		df4    = df[df['SUBCATEGORIA']==row['SUBCATEGORIA']]
		Sheet1.cell(row = 12 + index, column = 3).value = df4['ATIVIDADE'].nunique()
	Sheet1.title = 'Relatorio'
	df = df.drop(['Hora'], axis=1)
	df['DATA'] = df['DATA'].dt.strftime('%d/%m/%Y') 
	df['HORAS'] = df['HORAS'].dt.strftime('%H:%M')
	
	df.to_excel('temp.xlsx', sheet_name='Dados detalhados',index=False)
	temp_wb = openpyxl.load_workbook('temp.xlsx')
	ws2 = wb.create_sheet("Dados detalhados")
	for row in temp_wb['Dados detalhados']:
		for cell in row:
			ws2[cell.coordinate].value = cell.value
	'''
	new_file_name = 'relatorio_'+month_name+'.xlsx'
	wb.save(new_file_name)
	return new_file_name

# Criando o aplicativo Dash
app = dash.Dash(__name__)
server=app.server

# Definindo o layout do dashboard
app.layout = html.Div(children=[
    html.H1(children='Dashboard'),
    dcc.Upload(id='upload-data',children=html.Div(['Arraste e solte ou ',html.A('selecione arquivos')]),
              style={'width': '50%','height': '60px','lineHeight': '60px','borderWidth': '1px',
                     'borderStyle': 'dashed','borderRadius': '5px','textAlign': 'center','margin': '10px'},
              multiple=True),
              html.Div(id='output-data-upload'),
    html.Button('Baixar tabela', id='botao-download'),
    dcc.Download(id='download')
])

# Definindo o callback para gerar o arquivo Excel para download
@app.callback([Output('output-data-upload', 'children'),Output('download', 'data')], 
              Input('botao-download', 'n_clicks'), 
              [State('upload-data', 'contents'),State('upload-data', 'filename')])
def download_table(n_clicks, contents, filename):
    if n_clicks is None:
        return {},None
    #df = pd.DataFrame.from_dict(data)
    df = retorna_df(contents, filename)
    return {},dcc.send_file(preenche_modelo(12,2022,'victor',df), filename="tabela.xlsx")

if __name__ == '__main__':
    app.run_server(debug=True)
